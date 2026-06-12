"""
FreeEEG16 — Raw Data Recorder to Excel
=======================================
Records N seconds of raw 16-channel ADC data into RAM, then saves to .xlsx.

Requirements:
    pip install bleak numpy openpyxl

Usage:
    python freeeeg16_record.py --demo
    python freeeeg16_record.py --demo --duration 30
    python freeeeg16_record.py --out my_session.xlsx
"""

import asyncio
import argparse
import time
import threading

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────────
SERVICE_UUID        = "4fafc201-1fb5-459e-8fcc-c5c9c331914b"
DATA_CHAR_UUID      = "beb5483e-36e1-4688-b7f5-ea07361b26a8"
DEVICE_NAME         = "FreeEEG16"
PACKET_SIZE         = 51
HEADER_BYTE         = 0xA0
FOOTER_BYTE         = 0xC0
NUM_CHANNELS        = 16
SAMPLE_RATE         = 250
BLE_SCAN_TIMEOUT    = 30.0
BLE_CONNECT_TIMEOUT = 20.0

CHANNEL_LABELS = [f"CH{i+1}" for i in range(NUM_CHANNELS)]
CH_COLORS = [
    "E63946","F4A261","2A9D8F","457B9D",
    "A8DADC","E9C46A","7FC97F","8ECAE6",
    "FF6B6B","FFD166","06D6A0","118AB2",
    "FB8500","8338EC","3A86FF","FF006E",
]

# ── Packet parser ──────────────────────────────────────────────────────────────
def parse_int24_be(b0, b1, b2):
    val = (b0 << 16) | (b1 << 8) | b2
    return val - 0x1000000 if val & 0x800000 else val

def parse_packet(data: bytes):
    if len(data) != PACKET_SIZE or data[0] != HEADER_BYTE or data[50] != FOOTER_BYTE:
        return None
    channels = []
    for i in range(8):
        b = 2 + i * 3
        channels.append(parse_int24_be(data[b], data[b+1], data[b+2]))
    for i in range(8):
        b = 26 + i * 3
        channels.append(parse_int24_be(data[b], data[b+1], data[b+2]))
    return {"counter": data[1], "channels": channels}

# ── Recorder — stores everything in RAM as a numpy array ──────────────────────
class Recorder:
    def __init__(self, duration_sec: float):
        self.duration    = duration_sec
        self.target      = int(duration_sec * SAMPLE_RATE)
        # Pre-allocate: columns = [elapsed, counter, ch0..ch15]
        self._buf        = np.zeros((self.target, 2 + NUM_CHANNELS), dtype=np.float64)
        self._idx        = 0
        self._start_time = None
        self.done        = threading.Event()

    def push(self, channels, counter, wall_time):
        if self.done.is_set():
            return
        if self._start_time is None:
            self._start_time = wall_time
        if self._idx >= self.target:
            self.done.set()
            return
        elapsed = wall_time - self._start_time
        self._buf[self._idx, 0] = elapsed
        self._buf[self._idx, 1] = counter
        self._buf[self._idx, 2:] = channels
        self._idx += 1
        if self._idx >= self.target:
            self.done.set()

    @property
    def data(self):
        return self._buf[:self._idx]

    @property
    def sample_count(self):
        return self._idx

# ── BLE ────────────────────────────────────────────────────────────────────────
async def _find_device():
    from bleak import BleakScanner
    dev = await BleakScanner.find_device_by_filter(
        lambda d, adv: SERVICE_UUID.lower() in
                       [str(u).lower() for u in adv.service_uuids],
        timeout=BLE_SCAN_TIMEOUT,
    )
    if dev:
        return dev
    return await BleakScanner.find_device_by_name(DEVICE_NAME, timeout=BLE_SCAN_TIMEOUT)

async def _ble_record(recorder: Recorder):
    from bleak import BleakClient
    device = await _find_device()
    if device is None:
        recorder.done.set()
        return

    def on_data(_h, data: bytearray):
        r = parse_packet(bytes(data))
        if r:
            recorder.push(r["channels"], r["counter"], time.time())

    async with BleakClient(device, timeout=BLE_CONNECT_TIMEOUT) as client:
        if hasattr(client, "request_mtu"):
            try:
                await client.request_mtu(100)
            except Exception:
                pass
        await client.start_notify(DATA_CHAR_UUID, on_data)
        await asyncio.get_event_loop().run_in_executor(None, recorder.done.wait)
        await client.stop_notify(DATA_CHAR_UUID)

def ble_record(recorder: Recorder):
    asyncio.run(_ble_record(recorder))

# ── Demo generator ─────────────────────────────────────────────────────────────
def demo_record(recorder: Recorder):
    freqs  = np.array([1,2,4,8,10,13,15,20,25,30,35,40,45,50,55,60], dtype=np.float64)
    t0     = time.time()
    counter = 0

    while not recorder.done.is_set():
        now      = time.time()
        expected = min(int((now - t0) * SAMPLE_RATE), recorder.target)

        while counter < expected:
            t   = counter / SAMPLE_RATE
            ch  = (np.sin(2 * np.pi * freqs * t) * 500_000
                   + np.random.randn(NUM_CHANNELS) * 8000).astype(int).tolist()
            recorder.push(ch, counter & 0xFF, t0 + t)
            counter += 1
            if recorder.done.is_set():
                return

        time.sleep(0.005)

# ── Excel writer — writes from RAM buffer ──────────────────────────────────────
def save_excel(recorder: Recorder, path: str):
    data = recorder.data          # numpy array shape (N, 18)
    n    = len(data)
    if n == 0:
        return

    wb = openpyxl.Workbook(write_only=True)   # write_only = much faster

    thin   = Side(style="thin", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: Raw data ──────────────────────────────────────────────────
    ws = wb.create_sheet("Raw data")

    # Header
    headers    = ["Time (s)", "Counter"] + CHANNEL_LABELS
    color_list = [None, None] + CH_COLORS
    header_cells = []
    for hdr, raw_col in zip(headers, color_list):
        cell = openpyxl.cell.WriteOnlyCell(ws, value=hdr)
        cell.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        cell.fill      = PatternFill("solid",
                                     start_color=raw_col if raw_col else "2C2C2C")
        header_cells.append(cell)
    ws.append(header_cells)

    # Data rows — build from numpy, no per-cell font objects shared is fine
    data_font_even = Font(name="Arial", size=8)
    data_font_odd  = Font(name="Arial", size=8)
    fill_even      = PatternFill("solid", start_color="1A1A2E")
    fill_odd       = PatternFill("solid", start_color="16213E")
    num_align      = Alignment(horizontal="right")

    for r_idx, row in enumerate(data, start=2):
        fill = fill_even if r_idx % 2 == 0 else fill_odd
        cells = []
        for c_idx, val in enumerate(row):
            cell = openpyxl.cell.WriteOnlyCell(ws)
            cell.value     = round(float(val), 4) if c_idx == 0 else int(val)
            cell.font      = data_font_even
            cell.fill      = fill
            cell.border    = border
            cell.alignment = num_align
            cells.append(cell)
        ws.append(cells)

    # Stats — write_only can't use formulas referencing other rows easily,
    # so compute stats from numpy directly (faster anyway)
    ch_data    = data[:, 2:]      # shape (N, 16)
    stat_fill  = PatternFill("solid", start_color="0F3460")
    stat_font  = Font(name="Arial", bold=True, size=8, color="E0E0E0")
    stat_align = Alignment(horizontal="right")
    stat_fns   = [("MIN", np.min), ("MAX", np.max),
                  ("MEAN", np.mean), ("STD", np.std)]

    for label, fn in stat_fns:
        cells = []
        # Label cell
        lc = openpyxl.cell.WriteOnlyCell(ws, value=label)
        lc.font = stat_font; lc.fill = stat_fill
        lc.alignment = Alignment(horizontal="center"); lc.border = border
        cells.append(lc)
        # Empty counter cell
        ec = openpyxl.cell.WriteOnlyCell(ws, value="")
        ec.fill = stat_fill; ec.border = border
        cells.append(ec)
        # Per-channel stat values
        vals = fn(ch_data, axis=0)
        for v in vals:
            sc = openpyxl.cell.WriteOnlyCell(ws, value=round(float(v), 2))
            sc.font = stat_font; sc.fill = stat_fill
            sc.alignment = stat_align; sc.border = border
            cells.append(sc)
        ws.append(cells)

    # ── Sheet 2: Info ──────────────────────────────────────────────────────
    wi = wb.create_sheet("Info")
    hdr_fill  = PatternFill("solid", start_color="0F3460")
    hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    val_fill  = PatternFill("solid", start_color="16213E")
    val_font  = Font(name="Arial", size=10)
    val_font2 = Font(name="Arial", size=10, color="A8DADC")
    row_align = Alignment(horizontal="left", vertical="center")

    def info_cell(ws, value, font, fill):
        c = openpyxl.cell.WriteOnlyCell(ws, value=value)
        c.font = font; c.fill = fill; c.border = border; c.alignment = row_align
        return c

    wi.append([info_cell(wi, "Parameter", hdr_font, hdr_fill),
               info_cell(wi, "Value",     hdr_font, hdr_fill)])

    info_rows = [
        ("Device",         DEVICE_NAME),
        ("Duration (s)",   recorder.duration),
        ("Samples",        n),
        ("Sample rate",    f"{n / recorder.duration:.1f} sps"),
        ("Channels",       NUM_CHANNELS),
        ("Recorded at",    time.strftime("%Y-%m-%d %H:%M:%S")),
        ("ADC resolution", "24-bit signed"),
        ("Packet size",    f"{PACKET_SIZE} bytes"),
        ("BLE service",    SERVICE_UUID),
    ]
    for k, v in info_rows:
        wi.append([info_cell(wi, k, val_font,  val_fill),
                   info_cell(wi, v, val_font2, val_fill)])

    wb.save(path)

# ── Entry point ────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out",      default="eeg_recording.xlsx")
    parser.add_argument("--duration", type=float, default=10.0)
    parser.add_argument("--demo",     action="store_true")
    args = parser.parse_args()

    recorder = Recorder(duration_sec=args.duration)
    target   = recorder.target

    fn = demo_record if args.demo else ble_record
    worker = threading.Thread(target=fn, args=(recorder,), daemon=True)
    worker.start()
    recorder.done.wait()
    worker.join(timeout=2)

    save_excel(recorder, args.out)

if __name__ == "__main__":
    main()
